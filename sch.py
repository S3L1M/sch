import openpyxl as xl

# Open workbook & worksheet
wb = xl.load_workbook('sch.xlsx')
ws = wb['Sheet1']

# Add data
ws.append(['=NOW()'])
ws.cell(ws.max_row, 1).number_format = xl.styles.numbers.FORMAT_DATE_XLSX22

# Save & Close file
wb.save('sch.xlsx')
wb.close()
print("DONE !!!")